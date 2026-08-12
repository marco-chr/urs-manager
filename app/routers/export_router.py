from fastapi import APIRouter, Request, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app import models, audit
from app.auth import get_current_user, can_edit_system
from app.utils import bump_minor_version
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

router = APIRouter()

REQ_TYPE_LABELS = {
    "functional": "Functional",
    "non_functional": "Non-Functional",
    "regulatory": "Regulatory",
    "interface": "Interface",
    "environmental": "Environmental",
    "ehs": "EHS",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GMP_FILL = PatternFill("solid", fgColor="E2EFDA")
MUST_FILL = PatternFill("solid", fgColor="FCE4D6")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP = Alignment(wrap_text=True, vertical="top")


def _user(request, db):
    try:
        return get_current_user(request, db)
    except Exception:
        return None


def _add_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def _style_row(ws, row_idx, req, alt):
    fill = ALT_FILL if alt else PatternFill()
    if req.gmp_flag:
        fill = GMP_FILL
    for cell in ws[row_idx]:
        cell.border = THIN
        cell.alignment = WRAP
        if fill:
            cell.fill = fill


@router.get("/systems/{system_id}/export/excel")
def export_excel(request: Request, system_id: int, db: Session = Depends(get_db)):
    user = _user(request, db)
    if not user:
        raise HTTPException(401)
    system = db.query(models.System).filter(models.System.id == system_id).first()
    if not system:
        raise HTTPException(404)

    sections = {s.id: s.name for s in db.query(models.Section).filter(models.Section.system_id == system_id).all()}
    reqs = (
        db.query(models.Requirement)
        .filter(
            models.Requirement.system_id == system_id,
            models.Requirement.enabled != False,
        )
        .order_by(models.Requirement.section_id, models.Requirement.req_id)
        .all()
    )

    wb = Workbook()

    # --- Cover sheet ---
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.column_dimensions["A"].width = 20
    ws_cover.column_dimensions["B"].width = 60
    cover_data = [
        ("System", system.name),
        ("Description", system.description or ""),
        ("Status", system.status.upper()),
        ("Exported by", user.full_name or user.username),
    ]
    for r, (k, v) in enumerate(cover_data, start=1):
        ws_cover.cell(r, 1, k).font = Font(bold=True)
        ws_cover.cell(r, 2, v)

    # --- All requirements sheet ---
    ws_all = wb.create_sheet("All Requirements")
    headers = ["Section", "Req ID", "Type", "Description", "Must Have", "GMP/GEP", "Note"]
    _add_header(ws_all, headers)
    ws_all.column_dimensions["A"].width = 22
    ws_all.column_dimensions["B"].width = 12
    ws_all.column_dimensions["C"].width = 16
    ws_all.column_dimensions["D"].width = 55
    ws_all.column_dimensions["E"].width = 12
    ws_all.column_dimensions["F"].width = 8
    ws_all.column_dimensions["G"].width = 35

    for i, req in enumerate(reqs):
        sec_name = sections.get(req.section_id, "—") if req.section_id else "—"
        row = [
            sec_name,
            req.req_id or "",
            REQ_TYPE_LABELS.get(req.req_type, req.req_type),
            req.description,
            "Must Have" if req.must_have else "Nice to Have",
            req.gmp_flag or "GEP",
            req.note or "",
        ]
        ws_all.append(row)
        _style_row(ws_all, i + 2, req, i % 2 == 1)

    # --- Per-type sheets ---
    type_groups: dict[str, list] = {}
    for req in reqs:
        type_groups.setdefault(req.req_type or "functional", []).append(req)

    for rtype, type_reqs in type_groups.items():
        ws = wb.create_sheet(REQ_TYPE_LABELS.get(rtype, rtype)[:31])
        _add_header(ws, headers)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 55
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 35
        for i, req in enumerate(type_reqs):
            sec_name = sections.get(req.section_id, "—") if req.section_id else "—"
            row = [
                sec_name,
                req.req_id or "",
                REQ_TYPE_LABELS.get(req.req_type, req.req_type),
                req.description,
                "Must Have" if req.must_have else "Nice to Have",
                req.gmp_flag or "GEP",
                req.note or "",
            ]
            ws.append(row)
            _style_row(ws, i + 2, req, i % 2 == 1)

    # --- Sections sheet ---
    ws_sec = wb.create_sheet("Sections")
    ws_sec.append(["#", "Section Name", "# Requirements"])
    for cell in ws_sec[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, (sec_id, sec_name) in enumerate(sections.items(), start=1):
        count = sum(1 for r in reqs if r.section_id == sec_id)
        ws_sec.append([i, sec_name, count])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"URS_{system.name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/systems/{system_id}/import/excel")
async def import_excel(
    request: Request,
    system_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    user = _user(request, db)
    if not user:
        raise HTTPException(401)
    if not can_edit_system(user, system_id, db):
        raise HTTPException(403)
    if not file.filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(400, detail="Only .xlsx files are supported")

    content = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, detail="Could not read the Excel file — make sure it is a valid .xlsx")

    ws = wb.active
    imported = 0
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        value = row[0]
        if value is None:
            continue
        description = str(value).strip()
        if not description:
            continue
        req = models.Requirement(
            system_id=system_id,
            description=description,
            req_type="functional",
            must_have=True,
            gmp_flag="GEP",
            enabled=True,
            created_by=user.id,
        )
        db.add(req)
        imported += 1

    if imported:
        bump_minor_version(db, system_id)
        db.commit()
        # Audit a single entry for the batch import
        audit.log(db, user, "requirements", system_id, "IMPORT",
                  new_value={"imported": imported, "source": file.filename},
                  system_id=system_id)

    wb.close()
    return {"imported": imported}
